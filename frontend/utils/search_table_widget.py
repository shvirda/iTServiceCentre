"""
SearchTableWidget - универсальный компонент для таблиц с поиском, фильтрацией и CRUD операциями
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QLabel, QSpinBox,
    QTableWidget, QTableWidgetItem, QMessageBox, QDialog, QFormLayout, QDateEdit,
    QFileDialog, QComboBox
)
from PyQt6.QtCore import Qt, QDate
import logging
import csv
import os

logger = logging.getLogger(__name__)

# Попытка импортировать openpyxl для Excel
excel_available = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    excel_available = True
except ImportError:
    logger.warning("openpyxl not available, Excel export will be limited")


class SearchTableWidget(QWidget):
    """Универсальный компонент для поиска и управления данными через таблицу"""
    
    def __init__(self, api_endpoint: str, columns: list, filters: list, api_client,
                 parent=None, allow_edit=True, allow_delete=True):
        """
        Инициализация SearchTableWidget
        
        Args:
            api_endpoint: API endpoint (напр. '/api/equipment')
            columns: Список конфигураций колонок [{'name': 'id', 'label': 'ID', 'type': 'int', 'editable': False}, ...]
            filters: Список фильтров [{'name': 'search', 'label': 'Поиск:', 'type': 'text'}, ...]
            api_client: APIClient объект
            parent: Родительский виджет
            allow_edit: Разрешить редактирование записей
            allow_delete: Разрешить удаление записей
        """
        super().__init__(parent)
        self.api_endpoint = api_endpoint
        self.columns = columns
        self.filters = filters
        self.api_client = api_client
        self.allow_edit = allow_edit
        self.allow_delete = allow_delete
        self.current_filters = {}
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        """Инициализация UI"""
        layout = QVBoxLayout()
        
        # Панель фильтров
        filters_layout = QHBoxLayout()
        filters_layout.addWidget(QLabel("Фильтры:"))
        
        self.filter_widgets = {}
        for filter_config in self.filters:
            label = QLabel(filter_config.get('label', ''))
            filters_layout.addWidget(label)
            
            if filter_config['type'] == 'text':
                widget = QLineEdit()
                widget.setPlaceholderText("Введите значение...")
                widget.textChanged.connect(self.on_filter_changed)
            elif filter_config['type'] == 'int':
                widget = QSpinBox()
                widget.setMinimum(0)
                widget.setMaximum(999999)
                widget.valueChanged.connect(self.on_filter_changed)
            else:
                widget = QLineEdit()
            
            self.filter_widgets[filter_config['name']] = widget
            filters_layout.addWidget(widget)
        
        filters_layout.addStretch()
        
        # Кнопка поиска
        search_btn = QPushButton("Поиск")
        search_btn.clicked.connect(self.load_data)
        filters_layout.addWidget(search_btn)
        
        # Кнопка очистки
        clear_btn = QPushButton("Очистить")
        clear_btn.clicked.connect(self.clear_filters)
        filters_layout.addWidget(clear_btn)
        
        layout.addLayout(filters_layout)
        
        # Таблица
        self.table = QTableWidget()
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels([col['label'] for col in self.columns])
        self.table.resizeColumnsToContents()
        layout.addWidget(self.table)
        
        # Кнопки действий
        actions_layout = QHBoxLayout()
        
        # Кнопка добавления
        add_btn = QPushButton("+ Добавить")
        add_btn.clicked.connect(self.add_row)
        actions_layout.addWidget(add_btn)
        
        if self.allow_edit:
            edit_btn = QPushButton("Редактировать")
            edit_btn.clicked.connect(self.edit_row)
            actions_layout.addWidget(edit_btn)
        
        if self.allow_delete:
            delete_btn = QPushButton("Удалить")
            delete_btn.clicked.connect(self.delete_row)
            actions_layout.addWidget(delete_btn)
        
        refresh_btn = QPushButton("Обновить")
        refresh_btn.clicked.connect(self.load_data)
        actions_layout.addWidget(refresh_btn)
        
        # Кнопки экспорта/импорта
        actions_layout.addSpacing(20)
        
        export_btn = QPushButton("Экспорт")
        export_btn.setMenu(self._create_export_menu())
        actions_layout.addWidget(export_btn)
        
        import_btn = QPushButton("Импорт")
        import_btn.setMenu(self._create_import_menu())
        actions_layout.addWidget(import_btn)
        
        actions_layout.addStretch()
        layout.addLayout(actions_layout)
        
        self.setLayout(layout)
    
    def _create_export_menu(self):
        """Создать меню экспорта"""
        from PyQt6.QtWidgets import QMenu
        
        menu = QMenu("Экспорт", self)
        
        csv_action = menu.addAction("Экспорт в CSV")
        csv_action.triggered.connect(self.export_to_csv)
        
        if excel_available:
            excel_action = menu.addAction("Экспорт в Excel")
            excel_action.triggered.connect(self.export_to_excel)
        
        return menu
    
    def _create_import_menu(self):
        """Создать меню импорта"""
        from PyQt6.QtWidgets import QMenu
        
        menu = QMenu("Импорт", self)
        
        csv_action = menu.addAction("Импорт из CSV")
        csv_action.triggered.connect(self.import_from_csv)
        
        if excel_available:
            excel_action = menu.addAction("Импорт из Excel")
            excel_action.triggered.connect(self.import_from_excel)
        
        return menu
    
    def export_to_csv(self):
        """Экспорт данных в CSV файл"""
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Предупреждение", "Нет данных для экспорта")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт в CSV",
            f"{self.api_endpoint.strip('/')}_export.csv",
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                
                # Заголовки
                headers = [col['label'] for col in self.columns]
                writer.writerow(headers)
                
                # Данные
                for row in range(self.table.rowCount()):
                    row_data = []
                    for col in range(self.table.columnCount()):
                        item = self.table.item(row, col)
                        row_data.append(item.text() if item else '')
                    writer.writerow(row_data)
            
            QMessageBox.information(self, "Успех", f"Данные экспортированы в: {file_path}")
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать: {str(e)}")
    
    def export_to_excel(self):
        """Экспорт данных в Excel файл"""
        if not excel_available:
            QMessageBox.warning(self, "Предупреждение", "Excel экспорт недоступен")
            return
        
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Предупреждение", "Нет данных для экспорта")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт в Excel",
            f"{self.api_endpoint.strip('/')}_export.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            
            # Заголовки
            headers = [col['label'] for col in self.columns]
            ws.append(headers)
            
            # Жирный шрифт для заголовков
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Данные
            for row in range(self.table.rowCount()):
                row_data = []
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    row_data.append(item.text() if item else '')
                ws.append(row_data)
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(file_path)
            QMessageBox.information(self, "Успех", f"Данные экспортированы в: {file_path}")
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось экспортировать: {str(e)}")
    
    def import_from_csv(self):
        """Импорт данных из CSV файла"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Импорт из CSV",
            "",
            "CSV Files (*.csv);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f, delimiter=';')
                
                # Пропускаем заголовки
                headers = next(reader)
                
                # Подтверждение
                reply = QMessageBox.question(
                    self,
                    "Подтверждение",
                    f"Найдено {sum(1 for _ in reader)} записей. Добавить все записи?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                )
                
                if reply != QMessageBox.StandardButton.Yes:
                    return
                
                # Возвращаемся в начало
                f.seek(0)
                next(reader)  # Пропускаем заголовки снова
                
                success_count = 0
                error_count = 0
                
                for row in reader:
                    if not row or all(not cell.strip() for cell in row):
                        continue
                    
                    # Создаем словарь данных
                    row_data = {}
                    for i, header in enumerate(headers):
                        if i < len(row) and header.strip():
                            # Находим соответствующее поле по имени
                            for col in self.columns:
                                if col['label'] == header.strip():
                                    row_data[col['name']] = row[i].strip()
                                    break
                    
                    if row_data:
                        success, response, error = self.api_client._make_request(
                            'POST', self.api_endpoint, data=row_data
                        )
                        if success:
                            success_count += 1
                        else:
                            error_count += 1
                            logger.error(f"Error importing row: {error}")
                
                QMessageBox.information(
                    self,
                    "Импорт завершен",
                    f"Успешно импортировано: {success_count}\nОшибок: {error_count}"
                )
                self.load_data()
                
        except Exception as e:
            logger.error(f"Error importing from CSV: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось импортировать: {str(e)}")
    
    def import_from_excel(self):
        """Импорт данных из Excel файла"""
        if not excel_available:
            QMessageBox.warning(self, "Предупреждение", "Excel импорт недоступен")
            return
        
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Импорт из Excel",
            "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Получаем заголовки
            headers = [cell.value for cell in ws[1]]
            
            # Подтверждение
            row_count = ws.max_row - 1  # минус заголовок
            reply = QMessageBox.question(
                self,
                "Подтверждение",
                f"Найдено {row_count} записей. Добавить все записи?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply != QMessageBox.StandardButton.Yes:
                return
            
            success_count = 0
            error_count = 0
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(not cell for cell in row):
                    continue
                
                # Создаем словарь данных
                row_data = {}
                for i, header in enumerate(headers):
                    if header and i < len(row):
                        # Находим соответствующее поле по имени
                        for col in self.columns:
                            if col['label'] == str(header):
                                value = str(row[i]) if row[i] is not None else ''
                                if value and value != 'None':
                                    row_data[col['name']] = value
                                break
                
                if row_data:
                    success, response, error = self.api_client._make_request(
                        'POST', self.api_endpoint, data=row_data
                    )
                    if success:
                        success_count += 1
                    else:
                        error_count += 1
                        logger.error(f"Error importing row: {error}")
            
            QMessageBox.information(
                self,
                "Импорт завершен",
                f"Успешно импортировано: {success_count}\nОшибок: {error_count}"
            )
            self.load_data()
            
        except Exception as e:
            logger.error(f"Error importing from Excel: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Не удалось импортировать: {str(e)}")
    
    def get_filter_params(self):
        """Получить параметры фильтров из виджетов"""
        params = {'limit': 50, 'offset': 0}
        
        for filter_name, widget in self.filter_widgets.items():
            if isinstance(widget, QLineEdit):
                value = widget.text().strip()
                if value:
                    params[filter_name] = value
            elif isinstance(widget, QSpinBox):
                value = widget.value()
                if value > 0:
                    params[filter_name] = value
        
        return params
    
    def load_data(self):
        """Загрузить данные из API"""
        try:
            params = self.get_filter_params()
            success, response, error = self.api_client.get_from_api(self.api_endpoint, **params)
            
            if not success:
                QMessageBox.warning(self, "Ошибка", f"Не удалось загрузить данные: {error}")
                return
            
            # Парсим ответ
            if isinstance(response, dict):
                data_list = response.get('data', [])
            else:
                data_list = response
            
            self.populate_table(data_list)
            
        except Exception as e:
            logger.error(f"Error loading data: {str(e)}")
            QMessageBox.critical(self, "Ошибка", f"Произошла ошибка: {str(e)}")
    
    def populate_table(self, data_list):
        """Заполнить таблицу данными"""
        self.table.setRowCount(len(data_list))
        
        for row_idx, item_data in enumerate(data_list):
            for col_idx, column in enumerate(self.columns):
                field_name = column['name']
                value = item_data.get(field_name, '')
                
                table_item = QTableWidgetItem(str(value))
                
                # Если колонка не редактируемая, делаем её неактивной
                if not column.get('editable', False):
                    table_item.setFlags(table_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                
                self.table.setItem(row_idx, col_idx, table_item)
        
        self.table.resizeColumnsToContents()
    
    def on_filter_changed(self):
        """Обработчик изменения фильтров (если нужна автозагрузка)"""
        pass
    
    def clear_filters(self):
        """Очистить фильтры"""
        for widget in self.filter_widgets.values():
            if isinstance(widget, QLineEdit):
                widget.clear()
            elif isinstance(widget, QSpinBox):
                widget.setValue(0)
        
        self.load_data()
    
    def edit_row(self):
        """Редактировать выбранную строку"""
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите строку для редактирования")
            return
        
        # Получаем ID записи (обычно первая колонка)
        id_item = self.table.item(current_row, 0)
        if not id_item:
            return
        
        try:
            record_id = int(id_item.text())
        except:
            QMessageBox.warning(self, "Ошибка", "Не удалось получить ID записи")
            return
        
        # Собираем обновлённые данные из таблицы
        update_data = {}
        for col_idx, column in enumerate(self.columns):
            if column.get('editable', False):
                item = self.table.item(current_row, col_idx)
                if item:
                    value = item.text()
                    
                    # Конвертируем типы
                    if column['type'] == 'int':
                        try:
                            value = int(value)
                        except:
                            pass
                    elif column['type'] == 'float':
                        try:
                            value = float(value)
                        except:
                            pass
                    
                    update_data[column['name']] = value
        
        if not update_data:
            QMessageBox.warning(self, "Предупреждение", "Нет изменений для сохранения")
            return
        
        # Отправляем обновление на API
        endpoint = f"{self.api_endpoint}/{record_id}"
        success, response, error = self.api_client._make_request('PUT', endpoint, data=update_data)
        
        if success:
            QMessageBox.information(self, "Успех", "Запись успешно обновлена")
            self.load_data()
        else:
            QMessageBox.critical(self, "Ошибка", f"Не удалось обновить запись: {error}")
    
    def delete_row(self):
        """Удалить выбранную строку"""
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите строку для удаления")
            return
        
        id_item = self.table.item(current_row, 0)
        if not id_item:
            return
        
        try:
            record_id = int(id_item.text())
        except:
            QMessageBox.warning(self, "Ошибка", "Не удалось получить ID записи")
            return
        
        # Запрос подтверждения
        reply = QMessageBox.question(
            self,
            "Подтверждение",
            "Вы уверены, что хотите удалить эту запись?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Отправляем запрос на удаление
        endpoint = f"{self.api_endpoint}/{record_id}"
        success, response, error = self.api_client._make_request('DELETE', endpoint)
        
        if success:
            QMessageBox.information(self, "Успех", "Запись успешно удалена")
            self.load_data()
        else:
            QMessageBox.critical(self, "Ошибка", f"Не удалось удалить запись: {error}")
    
    def add_row(self):
        """Добавить новую запись через диалог"""
        from PyQt6.QtWidgets import QDialog, QFormLayout, QLineEdit, QTextEdit, QComboBox, QDialogButtonBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Добавить запись")
        dialog.setMinimumWidth(400)
        
        layout = QFormLayout()
        
        # Создаём поля для ввода на основе колонок
        input_widgets = {}
        for column in self.columns:
            if column.get('editable', True) and column['name'] != 'id':
                field_name = column['name']
                field_type = column.get('type', 'str')
                
                if field_type == 'text' or field_type == 'str':
                    widget = QLineEdit()
                    widget.setPlaceholderText(f"Введите {column['label'].lower()}...")
                elif field_type == 'int':
                    widget = QSpinBox()
                    widget.setMinimum(0)
                    widget.setMaximum(999999)
                elif field_type == 'float':
                    from PyQt6.QtWidgets import QDoubleSpinBox
                    widget = QDoubleSpinBox()
                    widget.setMinimum(0)
                    widget.setMaximum(999999.99)
                    widget.setDecimals(2)
                else:
                    widget = QLineEdit()
                    widget.setPlaceholderText(f"Введите {column['label'].lower()}...")
                
                input_widgets[field_name] = widget
                layout.addRow(f"{column['label']}:", widget)
        
        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        dialog.setLayout(layout)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Собираем данные
            new_data = {}
            for field_name, widget in input_widgets.items():
                if isinstance(widget, QLineEdit):
                    value = widget.text().strip()
                elif isinstance(widget, (QSpinBox, QDoubleSpinBox)):
                    value = widget.value()
                    if value == 0:
                        continue
                else:
                    continue
                
                if value:
                    new_data[field_name] = value
            
            if not new_data:
                QMessageBox.warning(self, "Предупреждение", "Введите хотя бы одно поле")
                return
            
            # Отправляем данные на API
            success, response, error = self.api_client._make_request('POST', self.api_endpoint, data=new_data)
            
            if success:
                QMessageBox.information(self, "Успех", "Запись успешно добавлена")
                self.load_data()
            else:
                QMessageBox.critical(self, "Ошибка", f"Не удалось добавить запись: {error}")
